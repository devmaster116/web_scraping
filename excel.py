from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
import json
import requests
from bs4 import BeautifulSoup # type: ignore
from openpyxl import load_workbook
from detect import *
from web_scrapy import *

def url_connect(weburl):
    page = "error"
    try:
        page = requests.get(weburl)
    except:
        print(page.content)
    return page

def url_content_name_search(page):
    soup = BeautifulSoup(page.content, "html.parser")
    results = "Hello ERROR"
    Name = results.split(" ")
    try:
        results = soup.find("h4", class_="mb-1 fw-bold auto-truncate")
        Name = results.text.strip()
    except:
        print("error")
    total_name = Name.split(" ")
    return total_name

existing_file = 'sample.xlsx'
wb_obj = load_workbook(existing_file)
sheet_obj = wb_obj.active
A = sheet_obj.cell(row=1, column=1)
A.value = "UserName"
B = sheet_obj.cell(row=1, column=2)
B.value = "URL"

for page_num in range(1, 2):
    url = "https://www.realtor.ca/realtor-search-results#sort=11-A&v=location&page="+str(page_num)
    page_content = cloudflare_bypass(url)

    soup = BeautifulSoup(page_content,"html.parser")
    print(page_content)

    for div_tag in soup.find_all("div", class_="realtorCardTopLeft realtorCardImageCon"):
        href_attr = div_tag.find("a", href=True)
        href_url = href_attr['href']
        img_src_attr = href_attr.find("img", src=True)
        image_src = img_src_attr['src']
        gender = gender_detect(image_src)
        if gender == "Female":
            next_div = div_tag.find_next_sibling('div', id_='realtorCardWrapDetails')
            span_tag = next_div.find("span","realtorCardName")
            userName = span_tag.text().strip()
            print(gender)
            print(url)
            excell_data = [userName, href_url]
            sheet_obj.append(excell_data)
            wb_obj.save(existing_file)
        else:
            wb_obj.save(existing_file)